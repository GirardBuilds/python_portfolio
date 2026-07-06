'''
M1P5 — Excel Reader/Writer (Ch 14)
Read an existing Excel file, display the data in the terminal,
perform a simple calculation on it (sum or average a column),
then write the results back to a new sheet in the same file or a new file entirely.
'''
import openpyxl
from pathlib import Path
file_path = Path.home() / 'Documents' /'python_portfolio' / 'learning' / 'chapter_study' / 'ATBS_chapter_Follow_along' / 'example3.xlsx'
wb = openpyxl.load_workbook(file_path)
#print(type(wb))
#print(wb.sheetnames)
sheet = wb['Sheet1']
Num_save = []
for i in range(1, 8): # Go through every other row.
    print(i, sheet.cell(row=i, column=2).value,f"{'total sold'.center(20,'_')} = ",sheet.cell(row=i ,column=3).value)
    Num_save.append(sheet.cell(row=i ,column=3).value)
sheet = wb['Sheet2']
sheet['C1'] = sum(Num_save)
sheet['C2'] = sum(Num_save) / len(Num_save)
sheet['A1'] = 'Total Sold'
sheet['A2'] = 'Average amount sold'
#print(sheet['C8'].value)
wb.save(file_path)
#print(sum(Num_save))
print('saved to file')









