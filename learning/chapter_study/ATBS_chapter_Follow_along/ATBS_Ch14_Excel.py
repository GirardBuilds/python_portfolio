import openpyxl
from pathlib import Path
print(Path.cwd())
'''
#Opening a Workbook
import openpyxl
wb = openpyxl.load_workbook('example3.xlsx')
type(wb) #<class 'openpyxl.workbook.workbook.Workbook'>
'''
'''
#Getting Sheets from the Workbook
wb = openpyxl.load_workbook('example3.xlsx')
print(wb.sheetnames)# The workbook's sheets' names
sheet = wb['Sheet3']# Get a sheet from the workbook.
print(sheet)
print(type(sheet))
print(sheet.title)# Get the sheet's title as a string.
another_sheet = wb.active# Get the active sheet.
print(another_sheet)
'''
'''
#Getting Cells from the Sheets
wb = openpyxl.load_workbook('example3.xlsx')
sheet = wb['Sheet1'] # get sheet from workbook
sheet['A1'] # get a cell from the sheet
print('>',sheet['A1'])
print('>>', sheet['A1'].value) # get the value from the cell
c = sheet['B1'] # get another cell from the sheet
print('>>>', c.value)
# get the row column, and value from a cell
print('>>>>', f'Row {c.row}, Column {c.column} is {c.value}')
print('>>>>>', f'Cell {c.coordinate} is {c.value}')
print('>>>>>>', sheet['C1'].value)
print(''.rjust(40,'='))

print('>',sheet.cell(row=1, column=2))
print('>>',sheet.cell(row=1,column=2).value)
for i in range(1,8,2):#go through every other row
    print(i,sheet.cell(row=i, column=2).value)
'''
'''
import openpyxl
wb = openpyxl.load_workbook('example3.xlsx')
sheet = wb['Sheet1']
print('>',sheet.max_row) # get the highest row number.
print('>>',sheet.max_column)
'''
'''
#Converting Between Column Letters and Numbers
from openpyxl.utils import get_column_letter, colimn_index_from_string
get_column_letter(1) #'A'
get_column_letter(2) # 'B'
get_column_letter(27) #'AA'
get_column_letter(900) # 'AHP'
wb = openpyxl.load_workbook('example3.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column) #'C'
column_index_from_string('A') # Get A's number. # 1
column_index_from_string('AA') # 27
'''
'''
#Getting Rows and Columns
import openpyxl
wb = openpyxl.load_workbook('example3.xlsx')
sheet = wb['Sheet1']
sheet['A1':'C3']  # Get cells A1 to C3.
for row_of_cell_objects in sheet['A1':'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print('--- END OF ROW ---')
'''






































