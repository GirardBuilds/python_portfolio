rows - go from top down (numbers)

columns are from left to right (letters)
//////////////////////////////////////////////////////////////////////
'Opening and Navigating a Workbook
//////////////////////////////////////////////////////////////////////
What it is:
    The openpyxl module lets Python read and write .xlsx Excel files.
    You start by loading a workbook file into a Workbook object, then access individual sheets through it —
    either by name or by grabbing whichever sheet was active when the file was last saved.

Why it matters:
    Most data youll encounter in the real world lives in spreadsheets —
    being able to open and navigate them programmatically means you can process
    thousands of rows without touching Excel manually.

Functions/attributes at a glance:

openpyxl.load_workbook(filename) — opens an .xlsx file and returns a Workbook object

wb.sheetnames — list of all sheet name strings in the workbook

wb['SheetName'] — returns the Worksheet object for that sheet

wb.active — returns whichever sheet was active when the file was last saved

sheet.title — the sheets name as a string

Example:
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

wb.sheetnames          # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet1']   # get specific sheet
sheet.title            # 'Sheet1'
sheet = wb.active      # get the active sheet

Gotcha:
    wb.active returns the sheet that was active when the file was last saved in Excel —
    it doesn't necessarily mean the first sheet. Never assume it's Sheet1 without checking.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Accessing Cells and Their Values
//////////////////////////////////////////////////////////////////////
What it is:
    Once you have a Worksheet object, you can access any cell by its coordinate string ('A1')
    or by passing row and column integers to .cell(). Each cell is a Cell object with a
    .value attribute that holds whatever is stored there, and .row, .column, and .coordinate attributes for its position.

Why it matters:
    Cells are the fundamental unit of a spreadsheet — knowing how to read and navigate them is the basis for every automation task.

Attributes at a glance:

.value — the cells stored content (number, string, datetime, or None)

.row — integer row number

.column — integer column number

.coordinate — string like 'B1'

Example:
sheet['A1'].value           # reads cell A1
sheet['B1'].value           # reads cell B1

c = sheet['B1']
c.row        # 1
c.column     # 2
c.coordinate # 'B1'

# Access by row/column integers (useful in loops)
sheet.cell(row=1, column=2).value    # same as sheet['B1'].value

# Loop every other row in column B
for i in range(1, 8, 2):
    print(sheet.cell(row=i, column=2).value)

Gotcha:
    Column numbers in .cell() start at 1, not 0 — column=1 is column A.
    This is easy to mix up if youre used to Pythons zero-indexing.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Sheet Dimensions and Column Conversion
//////////////////////////////////////////////////////////////////////
What it is:
    Every Worksheet has .max_row and .max_column attributes that tell you the extent of the data —
    useful for building loops that cover all rows or columns.
    Since max_column returns an integer but Excel uses letters, openpyxl provides two utility functions to convert between them.

Why it matters:
    You rarely know in advance how many rows are in a spreadsheet —
    these tools let your code adapt to whatever data is there.

Functions at a glance:

sheet.max_row — highest row number that contains data (integer)

sheet.max_column — highest column number that contains data (integer)

get_column_letter(n) — converts column integer to letter string (27 → 'AA')

column_index_from_string('A') — converts column letter to integer ('AA' → 27)

Example:
from openpyxl.utils import get_column_letter, column_index_from_string

sheet.max_row       # 7
sheet.max_column    # 3

get_column_letter(1)    # 'A'
get_column_letter(27)   # 'AA'

column_index_from_string('A')    # 1
column_index_from_string('AA')   # 27

# Practical use: loop all rows in a spreadsheet
for row in range(1, sheet.max_row + 1):
    print(sheet.cell(row=row, column=1).value)

Gotcha:
    The range for looping all rows is range(1, sheet.max_row + 1) — forgetting the + 1 means you skip the last row.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Slicing Rows and Columns
//////////////////////////////////////////////////////////////////////
What it is:
    You can slice a rectangular range of cells using sheet['A1':'C3'] syntax,
    which returns a tuple of tuples — outer tuples are rows, inner tuples are the Cell objects in that row.
    You can also access all rows or all columns via sheet.rows and sheet.columns,
    which must be converted to a list before indexing.

Why it matters:
    Lets you iterate over a block of cells without knowing exact coordinates for every cell —
    useful for batch reading or processing a region of a spreadsheet.

Example:
# Iterate a rectangular slice
for row_of_cells in sheet['A1':'C3']:
    for cell in row_of_cells:
        print(cell.coordinate, cell.value)

# Access a whole column
col_b = list(sheet.columns)[1]   # index 1 = second column (B)
for cell in col_b:
    print(cell.value)

# Access a whole row
row_1 = list(sheet.rows)[0]     # index 0 = first row
for cell in row_1:
    print(cell.value)

Gotcha:
    sheet.rows and sheet.columns return generator objects —
    you must wrap them in list() before indexing with [n], or iterate over them directly in a for loop.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Creating and Saving Workbooks
//////////////////////////////////////////////////////////////////////
What it is:
    openpyxl.Workbook() creates a brand new blank workbook in memory with a single default sheet called 'Sheet'.
    Nothing is written to disk until you call .save() with a filename.
    You can rename sheets by assigning to .title, and add or delete sheets with .create_sheet() and del.

Why it matters:
    Lets your program generate spreadsheet output from scratch —
    useful for producing reports, exports, or any structured data output without needing Excel open.

Functions/methods at a glance:

openpyxl.Workbook() — creates a new blank workbook in memory

wb.save(filename) — writes the workbook to disk as an .xlsx file

wb.create_sheet(index, title) — adds a new sheet; index and title are optional

del wb['SheetName'] — removes a sheet from the workbook

sheet.title = 'NewName' — renames a sheet

Example:
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sales Data'

wb.create_sheet(title='Summary')             # adds at the end
wb.create_sheet(index=0, title='Cover')      # inserts at position 0

del wb['Sheet']                              # delete default blank sheet

wb.save('my_report.xlsx')                    # write to disk
Gotcha:
    .save() silently overwrites the file if it already exists —
    always save to a new filename rather than overwriting the original in case your code has a bug.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Writing Values and Formulas to Cells
//////////////////////////////////////////////////////////////////////
What it is:
    Writing to a cell works exactly like assigning to a dictionary key —
    sheet['A1'] = value. You can assign numbers, strings,
    or even Excel formula strings (starting with =).
    The formula is stored as-is; Excel calculates it when the file is opened.
    To read back a formulas calculated result,
    you must open the file with data_only=True after Excel has evaluated and saved it.

Why it matters:
    This is how your program outputs data into a spreadsheet —
    everything from single values to full formula-driven calculations.

Example:
sheet['A1'] = 'Hello'
sheet['B1'] = 42
sheet['C1'] = '=SUM(A1:B1)'     # stored as a formula string

wb.save('output.xlsx')

# Read back the calculated result (must open in Excel and save first)
wb2 = openpyxl.load_workbook('output.xlsx', data_only=True)
wb2.active['C1'].value          # 542 (if Excel already calculated it)

Gotcha:
    openpyxl cannot calculate formulas itself —
    data_only=True only works if the file was previously opened and saved in Excel
    (which stores the last calculated result).
    A freshly created file will return None for formula cells even with data_only=True.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Font Styling
//////////////////////////////////////////////////////////////////////
What it is:
    You can apply font formatting to cells by creating a Font object with the styling you want
    and assigning it to a cells .font attribute. Import Font from openpyxl.styles
    and pass keyword arguments for size, bold, italic, and font name.

Why it matters:
    Lets your program produce polished, readable spreadsheets —
    highlighting headers, flagging certain rows, or matching an existing report format
    without manual formatting in Excel.

Example:
from openpyxl.styles import Font
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

bold_header = Font(name='Calibri', bold=True, size=14)
sheet['A1'].font = bold_header
sheet['A1'] = 'Report Header'

italic_note = Font(italic=True, size=10)
sheet['A2'].font = italic_note
sheet['A2'] = 'Generated automatically'

wb.save('styled.xlsx')
Key Font() arguments: name (font name string), size (integer), bold (True/False), italic (True/False)

Gotcha:
    Assigning a Font object to .font applies all its properties at once —
    if you dont specify a property (like size), it uses the openpyxl default (11pt Calibri),
    which may override the sheets existing style.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Row/Column Dimensions — Height and Width
//////////////////////////////////////////////////////////////////////
What it is:
    Every worksheet has row_dimensions and column_dimensions attributes that act like dictionaries.
    You access a row by its number and set .height, or access a column by its letter and set .width.

Why it matters:
    Programmatically sizing rows and columns saves time when generating reports where content length varies
    and you need consistent formatting across many files.

Example:
sheet.row_dimensions[1].height = 40        # set row 1 height
sheet.column_dimensions['B'].width = 25    # set column B width
sheet.column_dimensions['C'].hidden = True # hide column C

Gotcha:
    Row dimensions are accessed by integer (row_dimensions[1])
    but column dimensions are accessed by letter string (column_dimensions['B']) —
    mixing these up causes a KeyError.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Merging and Unmerging Cells
//////////////////////////////////////////////////////////////////////
What it is:
    merge_cells() combines a rectangular group of cells into one single cell.
    After merging, you write content only to the top-left cell of the range —
    the rest are absorbed. unmerge_cells() splits them back apart.

Why it matters:
    Common for header rows that span multiple columns in formatted reports.

Example:
sheet.merge_cells('A1:D1')       # merge A1 through D1 into one cell
sheet['A1'] = 'Monthly Report'   # write to top-left of the merged range

sheet.unmerge_cells('A1:D1')     # split them back

wb.save('merged.xlsx')

Gotcha:
    After merging, only the top-left cell address holds the content —
    trying to write to sheet['B1'] within a merged range will not work as expected.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Freeze Panes
//////////////////////////////////////////////////////////////////////
What it is:
    Setting sheet.freeze_panes to a cell coordinate freezes all rows above
    that cell and all columns to the left of it, keeping them visible while the user scrolls.
    Set it to None or 'A1' to unfreeze everything.

Why it matters:
    Essential for large spreadsheets where column or row headers need to stay visible while scrolling through hundreds of rows.

Example:
sheet.freeze_panes = 'A2'    # freeze row 1 (header row stays visible)
sheet.freeze_panes = 'B1'    # freeze column A only
sheet.freeze_panes = 'C2'    # freeze row 1 AND columns A and B
sheet.freeze_panes = None    # unfreeze everything

Gotcha:
    The frozen area is everything above and to the left of the specified cell —
    the cell itself is not frozen. 'A2' freezes row 1, not row 2.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Charts
//////////////////////////////////////////////////////////////////////
What it is:
    openpyxl can generate charts from cell data by chaining four objects together:
    a Reference (the data range), a Series (wraps the reference with a label),
    a chart object (the chart type), and add_chart() to place it on the sheet.

Why it matters:
    Lets your program produce visual data summaries without needing to manually create charts in Excel.

Example:
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, 11):
    sheet['A' + str(i)] = i * i    # add some data

# Build the chart
ref = openpyxl.chart.Reference(sheet, 1, 1, 1, 10)   # (sheet, col, startrow, col, endrow)
series = openpyxl.chart.Series(ref, title='Squares')
chart = openpyxl.chart.BarChart()
chart.title = 'Squares Chart'
chart.append(series)
sheet.add_chart(chart, 'C2')       # top-left corner of chart placement

wb.save('chart.xlsx')
Other chart types: openpyxl.chart.LineChart(), openpyxl.chart.PieChart(), openpyxl.chart.ScatterChart()
Gotcha: The Reference() arguments go (sheet, min_col, min_row, max_col, max_row) — the order is col before row, which is the opposite of how most spreadsheet ranges are described.
----------------------------------------------------------------------
++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
----------------------------------------------------------------------
'Summary:
//////////////////////////////////////////////////////////////////////

Load: openpyxl.load_workbook() → Workbook → wb['Sheet'] → Worksheet → sheet['A1'] or sheet.cell(row, col) → Cell → .value

Create: openpyxl.Workbook() → assign values with sheet['A1'] = x → wb.save(filename)

Sheet management: wb.create_sheet() adds sheets; del wb['name'] removes them; .title renames

Always save to a new filename to preserve the original in case of bugs

max_row / max_column drive loops; use get_column_letter() / column_index_from_string() to switch between letter and integer representations

Style with Font() objects assigned to .font; size rows/columns via row_dimensions[n].height and column_dimensions['A'].width

Formulas are written as strings starting with =; data_only=True reads the last calculated result (must be saved by Excel first)

freeze_panes keeps headers visible while scrolling — everything above and left of the set cell is frozen
